import openpyxl, os

# reading Excel spreadsheets with openpyxl

os.chdir(r"C:\Users\micha\Documents\Training")

# opening Excel file
workbook = openpyxl.load_workbook("example.xlsx")

# looking for sheets names
sheets = workbook.sheetnames
print(sheets)

# getting Sheet1
sheet1 = workbook["Sheet1"]

# getting A1 cell object
cellA1 = sheet1["A1"].value

# getting B2 cell object
cellB2 = sheet1["B2"].value

# printing cell
print(cellA1)
print(cellB2)

# getting column name: B3
print(str(sheet1.cell(row=3, column=2)))

# getting names from 2nd column
for i in range(1, 8):
    print(i, sheet1.cell(row=i, column=2).value)


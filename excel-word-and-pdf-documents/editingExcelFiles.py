import openpyxl, os

# creating new workbook
wb = openpyxl.Workbook()


def savingFile():
    wb.save("my_workbook.xlsx")


# Sheet
print(wb.sheetnames)

# sheet assigned
sheet = wb["Sheet"]

# is A1 cell blank?
print(sheet["A1"].value == None)

# writing values into the workbook
sheet["A1"] = 42
sheet["A2"] = "Hello!"

# saving file
savingFile()

# changing current directory
os.chdir(r"C:\Users\micha\Documents\Training")

# creating new sheet
sheet2 = wb.create_sheet()

# changing sheet2 title
sheet2.title = "My second sheet"

# saving file
savingFile()

# creating new sheet as the first in the workbook
wb.create_sheet(index=0, title="My first sheet")

# saving file
savingFile()

# getting sheet names
print(wb.sheetnames)
